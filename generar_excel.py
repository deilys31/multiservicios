"""
generar_excel.py
────────────────
Genera el archivo productos.xlsx de ejemplo con 5 productos ficticios.

Requisito:
    pip install openpyxl

Uso:
    python generar_excel.py

El archivo se crea en el mismo directorio donde se ejecuta el script.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Falta la librería openpyxl.")
    print("   Instálala con:  pip install openpyxl")
    raise SystemExit(1)

import os

# ── Datos de los productos de ejemplo ────────────────────────────
PRODUCTOS = [
    {
        "nombre":      "Cargador USB-C 65W",
        "precio":      "$24.99",
        "imagen":      "cargador-usb.svg",
        "descripcion": "Carga rápida compatible con laptops y celulares",
    },
    {
        "nombre":      "Audífonos Bluetooth",
        "precio":      "$39.99",
        "imagen":      "audifonos-bt.svg",
        "descripcion": "Sonido HD, 20h de batería, cancelación de ruido",
    },
    {
        "nombre":      "Cable HDMI 2m 4K",
        "precio":      "$12.99",
        "imagen":      "cable-hdmi.svg",
        "descripcion": "Resolución 4K UHD, conectores dorados resistentes",
    },
    {
        "nombre":      "Powerbank 20000mAh",
        "precio":      "$29.99",
        "imagen":      "powerbank.svg",
        "descripcion": "Carga 3 dispositivos simultáneamente",
    },
    {
        "nombre":      "Hub USB 7 Puertos",
        "precio":      "$18.99",
        "imagen":      "hub-usb.svg",
        "descripcion": "Compatible USB 3.0 y 2.0, con LED indicador",
    },
]

COLUMNAS = ["nombre", "precio", "imagen", "descripcion"]
ANCHOS   = [30, 12, 24, 50]        # ancho de columna en caracteres

# ── Estilos ───────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2563EB")   # azul primario
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color="C7D2FE"),
    right=Side(style="thin", color="C7D2FE"),
    top=Side(style="thin", color="C7D2FE"),
    bottom=Side(style="thin", color="C7D2FE"),
)

ROW_FILL_ODD  = PatternFill("solid", fgColor="EEF2FF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # ── Cabecera ─────────────────────────────────────────────────
    ws.append(COLUMNAS)
    for col_idx, _ in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = BORDER_THIN

    # ── Filas de datos ───────────────────────────────────────────
    for row_num, producto in enumerate(PRODUCTOS, start=2):
        fill = ROW_FILL_ODD if row_num % 2 != 0 else ROW_FILL_EVEN
        for col_idx, col_name in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=producto[col_name])
            cell.fill      = fill
            cell.border    = BORDER_THIN
            cell.alignment = Alignment(vertical="center")

    # ── Alto de fila ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for row in range(2, len(PRODUCTOS) + 2):
        ws.row_dimensions[row].height = 18

    # ── Ancho de columna ──────────────────────────────────────────
    for col_idx, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # ── Inmovilizar la fila de cabecera ───────────────────────────
    ws.freeze_panes = "A2"

    return wb


def main():
    destino = os.path.join(os.path.dirname(os.path.abspath(__file__)), "productos.xlsx")
    wb = build_workbook()
    wb.save(destino)
    print(f"✅ Archivo generado: {destino}")
    print("   Ahora sirve el sitio con Live Server y recarga la página.")


if __name__ == "__main__":
    main()
